#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import html as html_mod
import os
import re
import json
import time
import glob
import smtplib
import logging
import traceback
from pathlib import Path
from datetime import datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from multiprocessing import Pool, cpu_count
from typing import Dict, List, Union

import openai
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from tabulate import tabulate

from config import (
    EMAIL_FROM,
    EMAIL_TO,
    SMTP_SERVER,
    SMTP_USERNAME,
    SMTP_PASSWORD,
    SMTP_PORT,
    LOG_DIR,
    LLM_PROVIDER,
    LLM_MODEL,
    LLM_BASE_URL,
)
from api_clients import call_with_retries


# ----------------------------
# LLM abstraction layer
# ----------------------------

def llm_chat_completion(prompt: str, temperature: float = 0.0) -> str:
    """
    Send a chat completion request to the configured LLM provider.
    Returns the response content string.

    Supports: openai (default), anthropic, ollama, or any OpenAI-compatible endpoint.
    Configure via LLM_PROVIDER, LLM_MODEL, LLM_BASE_URL env vars.
    """
    if LLM_PROVIDER == "anthropic":
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=LLM_MODEL,
            max_tokens=4096,
            temperature=temperature,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text

    # OpenAI-compatible providers (openai, ollama, vllm, together, etc.)
    if LLM_BASE_URL:
        openai.api_base = LLM_BASE_URL
    response = openai.ChatCompletion.create(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
        n=1,
        temperature=temperature,
    )
    return response['choices'][0]['message']['content']

# ============================
# Logging
# ============================

from logging_config import setup_logging

logger = setup_logging(__name__, caller_file=__file__)

# Ensure LOG_DIR exists for artifacts/flags
Path(LOG_DIR).mkdir(parents=True, exist_ok=True)

# ----------------------------
# Helpers
# ----------------------------

def utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def utc_today_iso() -> str:
    return datetime.now(timezone.utc).date().isoformat()

# ----------------------------
# HTML report rendering
# ----------------------------

def generate_html_report_with_recommendations(report_entries, digest_summary, gpt_recommendations, market_regime="unknown", backtesting_html=""):
    """
    Generates an HTML report with summaries from the report entries, GPT-4o recommendations, and a plot of the top coins.
    """
    digest_items = ''.join(
        f'<li style="font-size:14px;line-height:1.6;">{item}</li>'
        for item in (digest_summary.get('surge_summary', []) if digest_summary else [])
    )
    tickers = ', '.join(digest_summary.get('tickers', [])) if digest_summary else 'N/A'

    digest_html = f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
        <tr>
            <td style="padding:20px;">
                <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Sundown Digest Summary</h3>
                <p style="font-size:14px;line-height:1.6;"><strong>Tickers Mentioned:</strong> {tickers}</p>
                <p style="font-size:14px;line-height:1.6;"><strong>News Summary:</strong></p>
                <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                    {digest_items}
                </ul>
            </td>
        </tr>
    </table>
    """

    # Market regime banner
    regime_colors = {"bull": "#d4edda", "bear": "#f8d7da", "sideways": "#fff3cd", "unknown": "#e2e3e5"}
    regime_labels = {"bull": "BULL MARKET", "bear": "BEAR MARKET", "sideways": "SIDEWAYS MARKET", "unknown": "REGIME UNKNOWN"}
    regime_msgs = {
        "bull": "Scoring accuracy is highest in bull markets. Signals are more reliable.",
        "bear": "Bear market detected. Exercise extra caution — scoring accuracy is reduced. Consider smaller position sizes.",
        "sideways": "Sideways/choppy market. Signals may be unreliable. Consider waiting for a clear trend.",
        "unknown": "Market regime could not be determined. Insufficient data for BTC moving averages.",
    }
    regime_banner = f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:{regime_colors.get(market_regime, '#e2e3e5')};margin-bottom:15px;">
        <tr>
            <td style="padding:12px 20px;">
                <strong style="font-size:15px;">{regime_labels.get(market_regime, 'UNKNOWN')}</strong>
                <span style="font-size:13px;margin-left:10px;">{regime_msgs.get(market_regime, '')}</span>
            </td>
        </tr>
    </table>
    """

    color_explanation = """
    <p style="font-size:14px;line-height:1.6;">
        <strong>Color Meaning:</strong><br>
        <span style="background-color:#d4edda;padding:2px 5px;border-radius:3px;">Green</span>: Indicates coins expected to surge or break out.<br>
        <span style="background-color:#ffe5b4;padding:2px 5px;border-radius:3px;">Orange</span>: Indicates coins not expected to surge.
    </p>
    """

    if not gpt_recommendations or not gpt_recommendations.get('recommendations'):
        recommendations_html = """
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    <p style="font-size:14px;line-height:1.6;">No coins are currently recommended for purchase based on the analysis.</p>
                </td>
            </tr>
        </table>
        """
        plot_html = ""
    else:
        sorted_recommendations = sorted(
            gpt_recommendations['recommendations'],
            key=lambda x: float(x.get('cumulative_score', 0) or 0),
            reverse=True
        )

        recommendation_items = ''
        for item in sorted_recommendations:
            matching_entry = next(
                (
                    e for e in report_entries
                    if re.sub(r'\s+', ' ', str(e.get("coin_name", ""))).strip().lower()
                    == re.sub(r'\s+', ' ', str(item.get("coin", ""))).strip().lower()
                ),
                None
            )

            if not matching_entry:
                logger.debug(
                    "No matching entry for GPT coin '%s'. Available sample: %s",
                    item.get('coin'),
                    [e.get('coin_name') for e in report_entries[:10]]
                )

            safe_coin_id = html_mod.escape(matching_entry['coin_id']) if matching_entry else ''
            coin_url = f"https://coinpaprika.com/coin/{safe_coin_id}/" if matching_entry else '#'
            # Derive slug for CoinGecko/CoinMarketCap from coin_id (e.g. "btc-bitcoin" -> "bitcoin")
            coin_slug = safe_coin_id.split('-', 1)[1] if safe_coin_id and '-' in safe_coin_id else ''
            coingecko_url = f"https://www.coingecko.com/en/coins/{coin_slug}" if coin_slug else '#'
            coinmarketcap_url = f"https://coinmarketcap.com/currencies/{coin_slug}/" if coin_slug else '#'
            cumulative_score_percentage = matching_entry.get('cumulative_score_percentage', 'N/A') if matching_entry else 'N/A'
            background_color = "#d4edda" if str(item.get("recommendation", '')).strip().lower() == "yes" else "#ffe5b4"
            coin_name = html_mod.escape(str(item.get("coin", "")).title())

            weighted_pct = matching_entry.get('weighted_score_percentage', 'N/A') if matching_entry else 'N/A'
            tp_target = matching_entry.get('take_profit_target_pct', 'N/A') if matching_entry else 'N/A'
            sl_target = matching_entry.get('stop_loss_target_pct', 'N/A') if matching_entry else 'N/A'
            rsi_info = html_mod.escape(str(matching_entry.get('rsi_explanation', ''))) if matching_entry else ''

            recommendation_items += f"""
            <li style="font-size:14px;line-height:1.6;margin-bottom:10px;background-color:{background_color};padding:10px;border-radius:5px;">
                <b><a href="{coin_url}" target="_blank" style="color:#264653;text-decoration:none;">{coin_name}</a></b> - {html_mod.escape(str(item.get("reason","")))}<br>
                <strong>Weighted Score:</strong> {weighted_pct}% &nbsp;|&nbsp;
                <strong>Equal Score:</strong> {cumulative_score_percentage}%<br>
                <strong>Exit Targets:</strong> Take Profit: +{tp_target}% &nbsp;|&nbsp; Stop Loss: -{sl_target}%<br>
                <span style="font-size:12px;color:#666;">{rsi_info}</span><br>
                <span style="font-size:13px;">
                    <a href="{coin_url}" target="_blank" style="color:#0077cc;text-decoration:none;">CoinPaprika</a>
                    &nbsp;|&nbsp;
                    <a href="{coingecko_url}" target="_blank" style="color:#0077cc;text-decoration:none;">CoinGecko</a>
                    &nbsp;|&nbsp;
                    <a href="{coinmarketcap_url}" target="_blank" style="color:#0077cc;text-decoration:none;">CoinMarketCap</a>
                </span>
            </li>
            """
        recommendations_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    {color_explanation}
                    <p style="font-size:14px;line-height:1.6;"><strong>Meaning of Cumulative Score Percentage:</strong> a higher percentage indicates a stronger potential based on historical data and analysis.</p>
                    <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                        {recommendation_items}
                    </ul>
                </td>
            </tr>
        </table>
        """

        # Embed the attached image in the HTML using CID
        cid = "top_coins_plot"
        plot_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;text-align:center;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Top Coins Cumulative Scores Over Time</h3>
                    <img src="cid:{cid}" alt="Top Coins Plot" style="width:100%;max-width:600px;height:auto;"/>
                </td>
            </tr>
        </table>
        """

    html_content = f"""
    <html>
    <body style="margin:0;padding:0;background-color:#f9f9f9;font-family:Arial,sans-serif;color:#333;">
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f9f9f9;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
                        <tr>
                            <td style="padding:20px;">
                                <h2 style="text-align:center;color:#264653;font-size:24px;margin:0;">Coin Analysis Report</h2>
                            </td>
                        </tr>
                        <tr>
                            <td>{regime_banner}</td>
                        </tr>
                        <tr>
                            <td>
                                {digest_html}
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {recommendations_html}
                            </td>
                        </tr>
                        {plot_html}
                        <tr>
                            <td>
                                {backtesting_html}
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:20px;">
                                <p style="text-align:center;color:#777;font-size:12px;margin:0;">Report generated on {utcnow_iso()}</p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """
    return html_content

# ----------------------------
# GPT batching (recommendations)
# ----------------------------

ROWS_PER_BATCH = 50

def gpt4o_summarize_batch(batch_df):
    try:
        # Replace NaN with None for serialization
        batch_df = batch_df.where(batch_df.notnull(), None)
        df_json = batch_df.to_dict(orient='records')
        dataset_json = json.dumps(df_json, indent=2).replace('%', '%%')
    except Exception as e:
        logger.error(f"Failed to serialize batch: {e}")
        return {"recommendations": []}

    prompt = f"""
You are provided with structured analysis data for multiple cryptocurrency coins.

Your task is to **summarize the data for each coin** using the available metrics and recommend "Yes" or "No".

---

### Interpret the following metrics:

- **liquidity_risk**: Indicates how easily the coin can be traded. Use "low", "medium", or "high".
- **price_change_score**:
    - 0 = No significant price changes
    - 1 = A notable increase in price in one time window (e.g., short-term)
    - 2 = Strong increases in two time windows
    - 3 = Consistent strong price momentum across short, medium, and long term
- **volume_change_score**:
    - 0 = No notable increase in trading activity
    - 1 = A moderate spike in trading volume in one time window
    - 2 = Sustained and strong trading activity over multiple periods
- **cumulative_score** (out of 20): Combined measure of price momentum, volume, sentiment, news, and technical indicators.
    - 0–5 = Weak overall signals
    - 6–11 = Moderate momentum or mixed indicators
    - 12+ = Strong breakout potential or sustained multi-signal alignment
- **trend_conflict**: 
    - "Yes" = The coin shows consistent monthly growth but lacks short-term support — may indicate early-stage breakout or lagging price action.

---

### Output Format (JSON Only):

```json
{{
"recommendations": [
    {{
    "coin": "Coin Name",
    "liquidity_risk": "Low/Medium/High",
    "cumulative_score": "Score Value",
    "recommendation": "Yes" or "No",
    "reason": "An intuitive summary explaining the coin’s price, volume, sentiment, trend, and any notable news."
    }}
]
}}
```

---

### Instructions:

- Use "recommendation": "Yes" if the coin shows strong potential:
    - High cumulative score (≥ 6)
    - Price/volume momentum (≥ 2)
    - Low/Medium liquidity risk
    - Positive or surge-indicating news sentiment
- Use "recommendation": "No" if:
    - Low cumulative score (≤ 2)
    - No recent momentum
    - Negative news or high liquidity risk
    - Or insufficient/notable data
- Treat trend_conflict = "Yes" as a possible early-stage breakout.
- Never quote raw scores without interpretation.
- Be fluent and confident. Avoid jargon or raw numbers.
- Don’t summarize the dataset as a whole.
- Include a short summary of any news.
- Return only valid JSON.

Here is the dataset:
{dataset_json}
"""

    try:
        content = call_with_retries(lambda: llm_chat_completion(prompt))

        match = re.search(r'```json(.*?)```', content, re.DOTALL)
        if match:
            json_data = json.loads(match.group(1).strip())
            return json_data
        else:
            logger.warning("No JSON block found in GPT response.")
            return {"recommendations": []}

    except Exception as e:
        logger.error(f"GPT-4o batch failed: {e}")
        logger.debug(traceback.format_exc())
        return {"recommendations": []}


def gpt4o_summarize_each_coin(df, rows_per_batch=ROWS_PER_BATCH, num_processes=None):
    """
    Process a DataFrame of coin metrics in parallel, batching by number of rows.
    """
    batches = [df.iloc[i:i + rows_per_batch] for i in range(0, len(df), rows_per_batch)]
    if num_processes is None:
        num_processes = min(cpu_count(), len(batches)) or 1

    logger.info(f"Processing {len(batches)} batches with {num_processes} processes.")
    with Pool(num_processes) as pool:
        results = pool.map(gpt4o_summarize_batch, batches)

    all_recommendations = []
    for result in results:
        if result and 'recommendations' in result:
            all_recommendations.extend(result['recommendations'])

    logger.info(f"Aggregated {len(all_recommendations)} recommendations from GPT.")
    return {"recommendations": all_recommendations}

def gpt4o_analyze_and_recommend(df):
    """
    Uses GPT-4o to analyze the final results DataFrame and provide structured recommendations for coin purchases.
    """
    try:
        df_json = df.to_dict(orient='records')
        dataset_json = json.dumps(df_json, indent=2).replace('%', '%%')
    except Exception as e:
        logger.error(f"Failed to serialize df_json: {e}")
        dataset_json = "{}"

    prompt = f"""
You are provided with structured analysis data for multiple cryptocurrency coins. Your task is to **evaluate each coin individually** and decide if it should be considered for **purchase based on the potential for a breakout or surge in value**.

---

Please follow this **step-by-step reasoning process** for each coin:

### Step-by-Step Evaluation:

1. **Check data completeness and uniqueness**:
   - If the coin record lacks required metrics or is a duplicate (based on name or ID), skip it.

2. **Assess Liquidity Risk**:
   - Use the `liquidity_risk` metric to judge how easily the coin can be bought or sold.
   - High liquidity risk is a red flag unless strongly offset by other metrics.

3. **Evaluate Sentiment and Market Momentum**:
   - Analyze `sentiment_score`. A high score indicates bullish community outlook.
   - Cross-check with `volume` and `price change scores` to confirm rising interest.

4. **Analyze Cumulative Score**:
   - Use the `cumulative_score` as a key signal of overall strength.
   - Consider coins with high cumulative scores and strong support from sentiment/volume as breakout candidates.

5. **Apply Decision Criteria**:
   - Recommend "Yes" only if there is **strong and clear evidence** supporting a breakout potential.
   - Recommend "No" only if there is **clear and confident evidence against** a breakout.

6. **Generate Explanation**:
   - Clearly justify the recommendation using data.
   - Reference at least two specific metrics.

---

### Output Format (Structured JSON Only):

```json
{{
  "recommendations": [
    {{
      "coin": "Coin Name",
      "liquidity_risk": "Low/Medium/High",
      "cumulative_score": "Score Value",
      "recommendation": "Yes/No",
      "reason": "A fluent, specific, data-driven explanation referencing key metrics."
    }}
  ]
}}
```

Now, here is the dataset:
{dataset_json}
"""

    try:
        content = call_with_retries(lambda: llm_chat_completion(prompt))

        json_match = re.search(r'```json(.*?)```', content, re.DOTALL)
        if json_match:
            json_content = json_match.group(1).strip()
            parsed_data = json.loads(json_content)
            return parsed_data

        logger.debug("No JSON content found in the LLM response.")
        return {"recommendations": []}

    except Exception as e:
        logger.error(f"Failed to complete LLM analysis: {e}")
        logger.debug(traceback.format_exc())
        return {"recommendations": []}

# ----------------------------
# Email utilities
# ----------------------------

def send_failure_email():
    """
    Sends an email with the current results when the script encounters an error.
    Uses a daily flag file to avoid spamming.
    """
    today = utc_today_iso()
    flag_file = os.path.join(LOG_DIR, f"email_sent_{today}.flag")

    if os.path.exists(flag_file):
        logger.debug(f"Email already sent today ({today}). Skipping email.")
        return

    # Clean prior flags (in LOG_DIR)
    for file in glob.glob(os.path.join(LOG_DIR, "email_sent_*.flag")):
        try:
            os.remove(file)
            logger.debug(f"Deleted old flag file: {file}")
        except Exception as e:
            logger.debug(f"Failed to delete flag file {file}: {e}")

    if os.path.exists(os.path.join(LOG_DIR, 'coin_analysis_report.xlsx')):
        file_contents = "Results file exists: coin_analysis_report.xlsx"
    else:
        file_contents = "No data available, as the results file was not created."

    html_content = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; color: #333; }}
            h2 {{ color: #c0392b; }}
            p {{ font-size: 14px; color: #555; }}
            .content {{ background-color: #f9f9f9; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }}
            .content pre {{ background-color: #f4f4f4; border: 1px solid #ccc; padding: 10px; border-radius: 3px; }}
        </style>
    </head>
    <body>
        <h2>Failure in Weekly Coin Analysis Script</h2>
        <p>The script encountered an error. Below is the current status:</p>
        <div class="content"><pre>{file_contents}</pre></div>
        <p>Time: {utcnow_iso()}</p>
    </body>
    </html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = "Failure in Weekly Coin Analysis Script"
    msg['From'] = EMAIL_FROM
    msg['Bcc'] = EMAIL_TO

    part = MIMEText(html_content, 'html')
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            logger.debug("Connecting to SMTP for failure email…")
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            recipients = [e.strip() for e in EMAIL_TO.split(",")]
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        logger.debug("Failure email sent successfully.")

        with open(flag_file, 'w') as f:
            f.write("Email sent")

    except Exception as e:
        logger.debug(f"Failed to send failure email: {e}")
        logger.debug(traceback.format_exc())

def print_command_line_report(report_entries):
    """
    Prints a command-line report of the daily coin analysis (to logs).
    """
    df = pd.DataFrame(report_entries)
    logger.debug("Coin Analysis Report")
    logger.debug("" + tabulate(df, headers="keys", tablefmt="grid"))
    logger.debug(f"Report generated on {utcnow_iso()}")

# ----------------------------
# Digest summarization (GPT)
# ----------------------------

def gpt4o_summarize_digest_and_extract_tickers(digest_text):
    """
    Uses GPT-4o to summarize the Sundown Digest and extract key tickers.
    """
    prompt = f"""
    Analyze the following digest entries and provide the following:
    1) A concise summary in bullet points (no more than 250 words) of key news items likely to cause surges in the value of the mentioned coins. 
    2) List the relevant cryptocurrency tickers beside each news item. Ensure there is no duplication.

    Text:
    {digest_text}

    Respond **only** in JSON format with 'surge_summary' and 'tickers' as keys. Ensure the tickers are in alphabetical order and there are no duplicate tickers.
    """

    try:
        response_content = call_with_retries(lambda: llm_chat_completion(prompt)).strip()
        json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
            try:
                return json.loads(json_str)
            except json.JSONDecodeError:
                logger.debug(f"Failed to decode JSON: {json_str}")
                return {"surge_summary": [], "tickers": []}
        else:
            logger.debug(f"No JSON found in the response: {response_content}")
            return {"surge_summary": [], "tickers": []}

    except Exception as e:
        if "rate limit" in str(e).lower() or "429" in str(e):
            logger.warning(f"Rate limit reached: {e}. Waiting 60s before retry...")
            time.sleep(60)
            try:
                retry_content = call_with_retries(lambda: llm_chat_completion(prompt)).strip()
                json_match = re.search(r'\{.*\}', retry_content, re.DOTALL)
                if json_match:
                    return json.loads(json_match.group(0))
            except Exception as retry_err:
                logger.error(f"Retry also failed: {retry_err}")
            return {"surge_summary": [], "tickers": []}

        logger.debug(f"An error occurred while summarizing the digest and extracting tickers: {e}")
        logger.debug(traceback.format_exc())
        return {"surge_summary": [], "tickers": []}

def summarize_sundown_digest(digest):
    """
    Summarizes the Sundown Digest content from the last three days and extracts tickers.
    """
    now = datetime.now(timezone.utc)
    three_days_ago = now - timedelta(days=3)

    digest_texts = []
    for entry in digest or []:
        try:
            entry_date = datetime.strptime(entry['date'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=timezone.utc)
            if entry_date >= three_days_ago:
                digest_texts.append(str(entry.get('text', '')))
        except Exception:
            continue

    combined_digest_text = " ".join(digest_texts)
    return gpt4o_summarize_digest_and_extract_tickers(combined_digest_text)

# ----------------------------
# Email sending with report
# ----------------------------

def send_email_with_report(html_content, attachment_path, plot_image_path=os.path.join(LOG_DIR, 'top_coins_plot.png'), recommendations=None):
    """
    Sends an email with an HTML report and attachments.
    """
    try:
        logger.debug(f"Preparing email with attachment: {attachment_path} and plot: {plot_image_path}")
        
        if not EMAIL_FROM or not EMAIL_TO:
            logger.error("EMAIL_FROM or EMAIL_TO is not set. Cannot send email.")
            return
        
        msg = MIMEMultipart('related')
        msg['Subject'] = "AI Generated Coin Analysis Report"
        msg['From'] = EMAIL_FROM
        msg['To'] = EMAIL_TO

        part = MIMEText(html_content, 'html')
        msg.attach(part)

        if recommendations and len(recommendations) > 0:
            logger.debug(f"Attaching plot image: {plot_image_path}")
            try:
                with open(plot_image_path, 'rb') as img_file:
                    mime_image = MIMEImage(img_file.read(), _subtype='png')
                    mime_image.add_header('Content-ID', '<top_coins_plot>')
                    mime_image.add_header('Content-Disposition', 'inline', filename=os.path.basename(plot_image_path))
                    msg.attach(mime_image)
            except Exception as e:
                logger.error(f"Error attaching plot image: {e}")
                logger.debug(traceback.format_exc())

        if os.path.exists(attachment_path):
            logger.debug(f"Attaching Excel file: {attachment_path}")
            try:
                with open(attachment_path, 'rb') as file:
                    part = MIMEApplication(file.read(), _subtype="xlsx")
                    part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(attachment_path))
                    msg.attach(part)
            except Exception as e:
                logger.error(f"Error attaching Excel file: {e}")
                logger.debug(traceback.format_exc())
        else:
            logger.warning(f"Attachment file not found: {attachment_path}")

        try:
            logger.debug(f"Connecting to SMTP server: {SMTP_SERVER}")
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
                recipients = [e.strip() for e in EMAIL_TO.split(",")]
                server.sendmail(EMAIL_FROM, recipients, msg.as_string())
            logger.debug("Email sent successfully.")
        except Exception as e:
            logger.error(f"Error sending email: {e}")
            logger.debug(traceback.format_exc())

    except Exception as e:
        logger.error(f"An error occurred in send_email_with_report: {e}")
        logger.debug(traceback.format_exc())

# ----------------------------
# Excel report
# ----------------------------

def save_report_to_excel(report_entries, filename=os.path.join(LOG_DIR, 'coin_analysis_report.xlsx')):
    """
    Saves the report entries to an Excel file with enhanced formatting and styling.
    """
    try:
        Path(LOG_DIR).mkdir(parents=True, exist_ok=True)

        df = pd.DataFrame(report_entries)
        df.to_excel(filename, index=False)
        logger.info(f"Initial report written to {filename}")

        workbook = load_workbook(filename)
        sheet = workbook.active

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        cell_font = Font(name="Arial", size=10)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Header styling + width
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                cell.border = thin_border
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Body styling + width
        for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logger.error(f"Error processing cell {cell.coordinate}: {e}")
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        sheet.freeze_panes = "A2"

        try:
            workbook.save(filename)
            logger.info(f"Report saved to {filename} with enhanced formatting.")
        finally:
            workbook.close()

        return filename

    except Exception as e:
        logger.error(f"Error saving the report to Excel: {e}")
        logger.debug(traceback.format_exc())
        return filename
